"""
Prediction Tracker — Self-Learning Model Evaluator
====================================================
1. Save daily predictions before games start
2. Grade them against ESPN final scores after games end
3. Analyze systematic errors → save calibration offsets
4. Apply calibration offsets in future predictions (loaded by game_predictor.py)
5. Export full history to Excel with MODEL RECORD dashboard

Storage:
  data/prediction_history.json   — prediction + grade history
  data/model_calibration.json    — learned bias corrections
  data/prediction_report.xlsx    — Excel export
"""

from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

DATA_DIR         = Path(__file__).parent.parent / "data"
HISTORY_FILE     = DATA_DIR / "prediction_history.json"
PROPS_FILE       = DATA_DIR / "props_history.json"
CALIBRATION_FILE = DATA_DIR / "model_calibration.json"
EXCEL_PATH       = DATA_DIR / "prediction_report.xlsx"

# How many recent graded days to use when computing calibration
CALIBRATION_WINDOW = 30


# ─── internal helpers ─────────────────────────────────────────────────────────

def _load_history() -> dict:
    if HISTORY_FILE.exists():
        try:
            return json.loads(HISTORY_FILE.read_text())
        except Exception:
            return {}
    return {}


def _save_history(history: dict) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    HISTORY_FILE.write_text(json.dumps(history, indent=2, default=str))


def _load_props_history() -> dict:
    if PROPS_FILE.exists():
        try:
            return json.loads(PROPS_FILE.read_text())
        except Exception:
            return {}
    return {}


def _save_props_history(history: dict) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    PROPS_FILE.write_text(json.dumps(history, indent=2, default=str))


def _pct(c: int, t: int) -> float:
    return round(c / t * 100, 1) if t else 0.0


# ─── public API ───────────────────────────────────────────────────────────────

def save_daily_predictions(date: str, games: list[dict]) -> None:
    """
    Store today's model predictions before games start.
    Each game dict must contain: game_id, home, away, game_time, predictions.
    """
    history = _load_history()

    # Never overwrite an already-graded record
    if date in history and history[date].get("graded_at"):
        return

    history[date] = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "graded_at":    None,
        "games": [
            {
                "game_id":    g.get("game_id", f"{g.get('away','?')}@{g.get('home','?')}"),
                "home":       g.get("home", ""),
                "away":       g.get("away", ""),
                "game_time":  g.get("game_time", ""),
                "predictions": g.get("predictions", {}),
                "actuals":    {},
                "grades":     {},
                "error_analysis": {},
            }
            for g in games
        ],
        "summary": {},
    }
    _save_history(history)
    print(f"[Tracker] Saved {len(games)} predictions for {date}")


def grade_predictions(date: str) -> dict:
    """
    Fetch final scores from ESPN, grade predictions, add error analysis,
    then trigger calibration update. Safe to call even if already graded.
    """
    history = _load_history()
    if date not in history:
        print(f"[Tracker] No predictions found for {date}")
        return {}

    record = history[date]
    if record.get("graded_at"):
        print(f"[Tracker] {date} already graded")
        return record.get("summary", {})

    actuals = _fetch_actuals(date)
    if not actuals:
        print(f"[Tracker] No final scores available yet for {date}")
        return {}

    ml_correct = ml_total = 0
    sp_correct = sp_total = 0
    ou_correct = ou_total = 0

    for game in record["games"]:
        result = (
            actuals.get(game["game_id"])
            or actuals.get(f"{game['away']}@{game['home']}")
        )
        if not result:
            continue

        hs  = result.get("home_score")
        as_ = result.get("away_score")
        if hs is None or as_ is None:
            continue

        game["actuals"] = {
            "home_score": hs,
            "away_score": as_,
            "winner":     game["home"] if hs > as_ else game["away"],
            "total":      hs + as_,
            "margin":     hs - as_,
        }

        preds  = game.get("predictions", {})
        grades: dict[str, Any] = {}

        # ── Moneyline ─────────────────────────────────────────────────────
        if preds.get("winner_pick"):
            ml_total += 1
            correct = preds["winner_pick"] == game["actuals"]["winner"]
            grades["winner_correct"] = correct
            if correct:
                ml_correct += 1

        # ── Spread ────────────────────────────────────────────────────────
        if preds.get("spread_pick") and preds.get("spread_line") is not None:
            sp_total += 1
            actual_margin = hs - as_
            line = float(preds["spread_line"])
            covered = (
                actual_margin > (-line)
                if preds["spread_pick"] == "HOME"
                else (-actual_margin) > (-line)
            )
            grades["spread_correct"] = covered
            if covered:
                sp_correct += 1

        # ── Total ─────────────────────────────────────────────────────────
        if preds.get("total_pick") and preds.get("total_line") is not None:
            ou_total += 1
            total_actual = hs + as_
            total_line   = float(preds["total_line"])
            correct_ou = (
                (preds["total_pick"] == "OVER"  and total_actual > total_line)
                or (preds["total_pick"] == "UNDER" and total_actual < total_line)
            )
            grades["total_correct"] = correct_ou
            if correct_ou:
                ou_correct += 1

        # ── Error analysis for self-learning ──────────────────────────────
        error: dict[str, Any] = {}
        mh = preds.get("model_home_score")
        ma = preds.get("model_away_score")
        mt = preds.get("model_total")

        if mh is not None and ma is not None:
            error["home_score_error"] = round(float(mh) - hs, 1)   # + = over-proj
            error["away_score_error"] = round(float(ma) - as_, 1)
        if mt is not None:
            error["total_error"] = round(float(mt) - (hs + as_), 1) # + = over-proj total

        # Was winner prediction correct?
        if preds.get("winner_pick"):
            error["winner_correct"] = grades.get("winner_correct", False)
            # If wrong, note the actual winner and margin for analysis
            if not error["winner_correct"]:
                error["missed_winner_reason"] = (
                    f"Picked {preds['winner_pick']} but {game['actuals']['winner']} "
                    f"won by {abs(hs - as_)} pts"
                )

        # Total direction error
        if mt is not None and preds.get("total_pick"):
            direction_right = (
                (preds["total_pick"] == "OVER"  and (hs + as_) > float(preds.get("total_line", mt)))
                or (preds["total_pick"] == "UNDER" and (hs + as_) < float(preds.get("total_line", mt)))
            )
            if not direction_right:
                actual_total = hs + as_
                error["missed_total_reason"] = (
                    f"Picked {preds['total_pick']} {preds.get('total_line','?')} "
                    f"but actual was {actual_total} "
                    f"({'OVER' if actual_total > float(preds.get('total_line', 0)) else 'UNDER'})"
                )

        game["grades"]         = grades
        game["error_analysis"] = error

    summary = {
        "moneyline": {"correct": ml_correct, "total": ml_total, "pct": _pct(ml_correct, ml_total)},
        "spread":    {"correct": sp_correct, "total": sp_total, "pct": _pct(sp_correct, sp_total)},
        "total":     {"correct": ou_correct, "total": ou_total, "pct": _pct(ou_correct, ou_total)},
    }
    record["summary"]   = summary
    record["graded_at"] = datetime.now(timezone.utc).isoformat()
    _save_history(history)

    print(f"[Tracker] Graded {date}: "
          f"ML {summary['moneyline']['pct']}% | "
          f"Spread {summary['spread']['pct']}% | "
          f"O/U {summary['total']['pct']}%")

    # Auto-update calibration after every graded day
    analyze_and_calibrate()

    # Auto-export Excel so the file is always up to date
    try:
        export_to_excel()
    except Exception as _e:
        print(f"[Tracker] Excel export failed: {_e}")

    return summary


# ─── player props tracking ────────────────────────────────────────────────────

def save_daily_props(date: str, props: list[dict]) -> None:
    """
    Store today's prop predictions before games start.
    Each prop dict should have: player, team, opponent, stat, line, direction,
    avg, hit_rate, ev, is_lock.
    Safe to call multiple times — will not overwrite an already-graded day.
    """
    history = _load_props_history()

    # Never overwrite a graded record
    if date in history and history[date].get("graded_at"):
        return

    # Keep only the fields useful for grading and export
    KEEP = {"player", "team", "opponent", "stat", "line", "direction",
            "avg", "hit_rate", "ev", "is_lock", "game_matchup"}
    clean = []
    for p in props:
        entry = {k: p[k] for k in KEEP if k in p}
        entry["actual"]  = None   # filled in when graded
        entry["hit"]     = None
        clean.append(entry)

    history[date] = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "graded_at":    None,
        "props":        clean,
        "summary":      {},
    }
    _save_props_history(history)
    print(f"[PropsTracker] Saved {len(clean)} props for {date}")


def grade_props(date: str) -> dict:
    """
    Grade saved props against actual game logs.
    Looks up each player's actual stat in the Parquet game log for the given date.
    Returns a summary dict.
    """
    history = _load_props_history()
    if date not in history:
        print(f"[PropsTracker] No props saved for {date}")
        return {}

    record = history[date]
    if record.get("graded_at"):
        print(f"[PropsTracker] {date} already graded")
        return record.get("summary", {})

    # Load game logs for actual stats
    logs_df = _load_game_logs_for_date(date)
    if logs_df is None or logs_df.empty:
        print(f"[PropsTracker] No game log data available for {date}")
        return {}

    hit = miss = 0
    by_stat: dict[str, dict] = {}

    for prop in record["props"]:
        player   = prop["player"]
        stat     = prop["stat"]
        line     = prop.get("line")
        direction = prop.get("direction", "Over")

        if line is None:
            continue

        actual = _lookup_actual_stat(logs_df, player, stat, date)
        if actual is None:
            continue   # player didn't play or data not yet available

        prop["actual"] = actual
        is_hit = (actual >= line) if direction == "Over" else (actual <= line)
        prop["hit"] = is_hit

        if is_hit:
            hit += 1
        else:
            miss += 1

        # Per-stat breakdown
        s = by_stat.setdefault(stat, {"hit": 0, "miss": 0})
        if is_hit:
            s["hit"] += 1
        else:
            s["miss"] += 1

    total = hit + miss
    stat_summary = {
        s: {"hit": v["hit"], "total": v["hit"]+v["miss"],
            "pct": _pct(v["hit"], v["hit"]+v["miss"])}
        for s, v in by_stat.items()
    }
    summary = {
        "hit":        hit,
        "miss":       miss,
        "total":      total,
        "pct":        _pct(hit, total),
        "by_stat":    stat_summary,
    }
    record["summary"]   = summary
    record["graded_at"] = datetime.now(timezone.utc).isoformat()
    _save_props_history(history)

    print(f"[PropsTracker] Graded {date}: {hit}/{total} = {_pct(hit, total)}%")

    try:
        export_to_excel()
    except Exception as _e:
        print(f"[PropsTracker] Excel export failed: {_e}")

    return summary


def get_props_record() -> dict:
    """Aggregate all-time props record across every graded day."""
    history = _load_props_history()

    all_hit = all_total = 0
    by_stat: dict[str, dict] = {}
    recent_hits = recent_total = 0
    recent_dates = sorted(
        [d for d in history if history[d].get("graded_at")], reverse=True
    )[:7]

    for date in history:
        if not history[date].get("graded_at"):
            continue
        s = history[date].get("summary", {})
        all_hit   += s.get("hit", 0)
        all_total += s.get("total", 0)
        if date in recent_dates:
            recent_hits  += s.get("hit", 0)
            recent_total += s.get("total", 0)

        for stat, data in s.get("by_stat", {}).items():
            bs = by_stat.setdefault(stat, {"hit": 0, "total": 0})
            bs["hit"]   += data.get("hit", 0)
            bs["total"] += data.get("total", 0)

    by_stat_pct = {
        s: {"hit": v["hit"], "total": v["total"], "pct": _pct(v["hit"], v["total"])}
        for s, v in by_stat.items()
    }

    return {
        "hit":         all_hit,
        "miss":        all_total - all_hit,
        "total":       all_total,
        "pct":         _pct(all_hit, all_total),
        "recent_7d":   _pct(recent_hits, recent_total),
        "by_stat":     by_stat_pct,
        "days_graded": len([d for d in history if history[d].get("graded_at")]),
    }


def _load_game_logs_for_date(date: str):
    """Load player game logs from Parquet, filtered to the given date."""
    try:
        import pandas as pd
        parquet = DATA_DIR / "player_game_logs.parquet"
        csv_path = DATA_DIR / "player_game_logs.csv"

        if parquet.exists():
            df = pd.read_parquet(parquet)
        elif csv_path.exists():
            df = pd.read_csv(csv_path)
        else:
            return None

        if "_date" not in df.columns:
            df["_date"] = pd.to_datetime(
                df["GAME_DATE"], format="%b %d, %Y", errors="coerce"
            )

        target = pd.to_datetime(date)
        return df[df["_date"].dt.date == target.date()]
    except Exception as e:
        print(f"[PropsTracker] Could not load game logs: {e}")
        return None


def _lookup_actual_stat(df, player: str, stat: str, date: str):
    """
    Return the actual value of `stat` for `player` on `date`.
    Handles combo stats like 'PTS+REB+AST'.
    Returns None if not found.
    """
    try:
        import pandas as pd
        player_rows = df[df["PLAYER_NAME"] == player]
        if player_rows.empty:
            return None

        # Combo stat: sum the parts
        if "+" in stat:
            parts = stat.split("+")
            total = 0.0
            for part in parts:
                if part in player_rows.columns:
                    total += float(player_rows[part].iloc[0])
                else:
                    return None
            return round(total, 1)
        else:
            if stat not in player_rows.columns:
                return None
            return round(float(player_rows[stat].iloc[0]), 1)
    except Exception:
        return None


def analyze_and_calibrate() -> dict:
    """
    Analyze recent prediction errors and save calibration offsets.

    Looks at the last CALIBRATION_WINDOW graded days and computes:
      - total_bias:      avg (model_total - actual_total). Positive = model over-projects.
      - home_score_bias: avg (model_home - actual_home).
      - away_score_bias: avg (model_away - actual_away).
      - ml_home_bias:    fraction of time model picks home winner incorrectly.

    These are saved to data/model_calibration.json and applied by game_predictor.py.
    """
    history = _load_history()

    total_errors:      list[float] = []
    home_score_errors: list[float] = []
    away_score_errors: list[float] = []
    home_winner_picks  = 0
    home_winner_correct = 0

    sorted_dates = sorted(
        [d for d in history if history[d].get("graded_at")],
        reverse=True
    )[:CALIBRATION_WINDOW]

    for date in sorted_dates:
        for g in history[date].get("games", []):
            err = g.get("error_analysis", {})
            if "total_error" in err:
                total_errors.append(err["total_error"])
            if "home_score_error" in err:
                home_score_errors.append(err["home_score_error"])
            if "away_score_error" in err:
                away_score_errors.append(err["away_score_error"])
            preds = g.get("predictions", {})
            if preds.get("winner_pick") == g.get("home"):
                home_winner_picks += 1
                if g.get("grades", {}).get("winner_correct"):
                    home_winner_correct += 1

    def _avg(lst: list[float]) -> float:
        return round(sum(lst) / len(lst), 2) if lst else 0.0

    calibration = {
        "updated_at":       datetime.now(timezone.utc).isoformat(),
        "sample_days":      len(sorted_dates),
        "sample_games":     len(total_errors),
        # Positive bias = model projects too high → subtract this from future totals
        "total_bias":       _avg(total_errors),
        "home_score_bias":  _avg(home_score_errors),
        "away_score_bias":  _avg(away_score_errors),
        # If model is picking home teams more than they win, flag the bias
        "home_winner_pick_rate":    round(home_winner_picks / max(home_winner_picks + (len(total_errors) - home_winner_picks), 1), 3),
        "home_winner_accuracy":     _pct(home_winner_correct, home_winner_picks),
        "notes": _generate_calibration_notes(
            _avg(total_errors), _avg(home_score_errors), _avg(away_score_errors)
        ),
    }

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    CALIBRATION_FILE.write_text(json.dumps(calibration, indent=2))

    print(f"[Calibration] Updated from {len(sorted_dates)} days / {len(total_errors)} games:")
    print(f"  Total bias:      {calibration['total_bias']:+.1f} pts (model projects {abs(calibration['total_bias']):.1f} pts {'too high' if calibration['total_bias'] > 0 else 'too low'})")
    print(f"  Home score bias: {calibration['home_score_bias']:+.1f} pts")
    print(f"  Away score bias: {calibration['away_score_bias']:+.1f} pts")

    return calibration


def _generate_calibration_notes(total_bias: float, home_bias: float, away_bias: float) -> list[str]:
    """Plain-English explanations of what the model is getting wrong."""
    notes = []
    if abs(total_bias) >= 3:
        direction = "over-projecting" if total_bias > 0 else "under-projecting"
        notes.append(
            f"Model consistently {direction} game totals by {abs(total_bias):.1f} pts. "
            f"Correction applied to future total predictions."
        )
    if abs(home_bias) >= 3:
        direction = "over-projecting" if home_bias > 0 else "under-projecting"
        notes.append(
            f"Model {direction} home team scores by {abs(home_bias):.1f} pts on average."
        )
    if abs(away_bias) >= 3:
        direction = "over-projecting" if away_bias > 0 else "under-projecting"
        notes.append(
            f"Model {direction} away team scores by {abs(away_bias):.1f} pts on average."
        )
    if not notes:
        notes.append("No significant systematic bias detected in recent predictions.")
    return notes


def load_calibration() -> dict:
    """Load calibration offsets. Returns zero-offset defaults if no file exists yet."""
    if CALIBRATION_FILE.exists():
        try:
            return json.loads(CALIBRATION_FILE.read_text())
        except Exception:
            pass
    return {
        "total_bias": 0.0,
        "home_score_bias": 0.0,
        "away_score_bias": 0.0,
    }


def get_model_record() -> dict:
    """
    Aggregate all-time model record across every graded day.
    Returns totals for ML, Spread, O/U plus recent-form (last 7 days).
    """
    history = _load_history()

    all_ml_c = all_ml_t = 0
    all_sp_c = all_sp_t = 0
    all_ou_c = all_ou_t = 0

    recent_dates = sorted(
        [d for d in history if history[d].get("graded_at")],
        reverse=True
    )[:7]
    r_ml_c = r_ml_t = 0
    r_sp_c = r_sp_t = 0
    r_ou_c = r_ou_t = 0

    for date in history:
        if not history[date].get("graded_at"):
            continue
        s = history[date].get("summary", {})
        ml = s.get("moneyline", {})
        sp = s.get("spread", {})
        ou = s.get("total", {})
        all_ml_c += ml.get("correct", 0);  all_ml_t += ml.get("total", 0)
        all_sp_c += sp.get("correct", 0);  all_sp_t += sp.get("total", 0)
        all_ou_c += ou.get("correct", 0);  all_ou_t += ou.get("total", 0)
        if date in recent_dates:
            r_ml_c += ml.get("correct", 0); r_ml_t += ml.get("total", 0)
            r_sp_c += sp.get("correct", 0); r_sp_t += sp.get("total", 0)
            r_ou_c += ou.get("correct", 0); r_ou_t += ou.get("total", 0)

    all_c = all_ml_c + all_sp_c + all_ou_c
    all_t = all_ml_t + all_sp_t + all_ou_t

    return {
        "moneyline": {"correct": all_ml_c, "wrong": all_ml_t - all_ml_c, "total": all_ml_t, "pct": _pct(all_ml_c, all_ml_t)},
        "spread":    {"correct": all_sp_c, "wrong": all_sp_t - all_sp_c, "total": all_sp_t, "pct": _pct(all_sp_c, all_sp_t)},
        "total":     {"correct": all_ou_c, "wrong": all_ou_t - all_ou_c, "total": all_ou_t, "pct": _pct(all_ou_c, all_ou_t)},
        "overall":   {"correct": all_c,    "wrong": all_t - all_c,       "total": all_t,    "pct": _pct(all_c, all_t)},
        "recent_7d": {
            "moneyline": _pct(r_ml_c, r_ml_t),
            "spread":    _pct(r_sp_c, r_sp_t),
            "total":     _pct(r_ou_c, r_ou_t),
        },
        "days_graded": len([d for d in history if history[d].get("graded_at")]),
    }


def export_to_excel(output_path: str | None = None) -> str:
    """
    Export full prediction history to Excel with 5 sheets:
      1. Model Record  — all-time W/L record dashboard (shown first)
      2. Summary       — one row per date
      3. Moneyline     — per-game picks + results
      4. Spread        — per-game picks + results
      5. Total (O/U)   — per-game picks + results

    Green rows = correct prediction, Red rows = wrong.
    Returns the path to the written file.
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl is required. Run: pip install openpyxl")

    history = _load_history()
    if not history:
        raise ValueError("No prediction history found yet.")

    path = Path(output_path) if output_path else EXCEL_PATH
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Style helpers ──────────────────────────────────────────────────────────
    green       = PatternFill("solid", fgColor="C6EFCE")
    red         = PatternFill("solid", fgColor="FFC7CE")
    gold        = PatternFill("solid", fgColor="FFD700")
    navy        = PatternFill("solid", fgColor="1F4E79")
    dark_green  = PatternFill("solid", fgColor="375623")
    dark_red    = PatternFill("solid", fgColor="9C0006")

    bold_white  = Font(bold=True, color="FFFFFF")
    bold_black  = Font(bold=True, color="000000")
    bold_gold   = Font(bold=True, color="FFD700", size=14)
    center      = Alignment(horizontal="center", vertical="center")
    left        = Alignment(horizontal="left",   vertical="center")

    def _make_sheet(title: str, headers: list[str]) -> Any:
        ws = wb.create_sheet(title)
        ws.append(headers)
        for i, _ in enumerate(headers, 1):
            c = ws.cell(1, i)
            c.fill = navy; c.font = bold_white; c.alignment = center
        ws.row_dimensions[1].height = 20
        return ws

    def _col_widths(ws: Any, widths: list[int]) -> None:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _fill_row(ws: Any, correct: bool | None) -> None:
        if correct is None:
            return
        f = green if correct else red
        for col in range(1, ws.max_column + 1):
            ws.cell(ws.max_row, col).fill = f

    # ── Sheet 1: MODEL RECORD (dashboard) ─────────────────────────────────────
    ws_rec = wb.create_sheet("Model Record", 0)
    ws_rec.sheet_view.showGridLines = False
    record = get_model_record()
    calib  = load_calibration()

    def _rec_cell(row: int, col: int, value: Any,
                  fill=None, font=None, align=None) -> None:
        c = ws_rec.cell(row, col, value)
        if fill:  c.fill  = fill
        if font:  c.font  = font
        if align: c.alignment = align

    # Title
    ws_rec.merge_cells("A1:G1")
    _rec_cell(1, 1, "🏀  NBA PREDICTOR — MODEL RECORD",
              fill=navy, font=bold_gold, align=center)
    ws_rec.row_dimensions[1].height = 30

    # Updated at
    ws_rec.merge_cells("A2:G2")
    _rec_cell(2, 1, f"Last updated: {datetime.now().strftime('%A %B %d, %Y  %I:%M %p')}",
              font=Font(italic=True, color="595959"), align=center)

    # Section header
    ws_rec.merge_cells("A4:G4")
    _rec_cell(4, 1, "ALL-TIME RECORD", fill=PatternFill("solid", fgColor="2E4057"),
              font=Font(bold=True, color="FFFFFF", size=11), align=center)

    # Column headers
    headers = ["Category", "Correct ✓", "Wrong ✗", "Total", "Win Rate", "Status", ""]
    for col, h in enumerate(headers, 1):
        _rec_cell(5, col, h, fill=navy, font=bold_white, align=center)

    row = 6
    categories = [
        ("Moneyline",  record["moneyline"]),
        ("Spread",     record["spread"]),
        ("Over/Under", record["total"]),
        ("OVERALL",    record["overall"]),
    ]
    for label, data in categories:
        pct  = data["pct"]
        status = "🔥 HOT" if pct >= 60 else ("✅ OK" if pct >= 50 else "⚠️ COLD")
        is_overall = (label == "OVERALL")
        row_fill = PatternFill("solid", fgColor="D9E1F2") if not is_overall else gold
        row_font = Font(bold=True) if is_overall else None
        for col, val in enumerate(
            [label, data["correct"], data["wrong"], data["total"], f"{pct}%", status, ""], 1
        ):
            c = ws_rec.cell(row, col, val)
            c.fill = row_fill
            if row_font: c.font = row_font
            c.alignment = center
        row += 1

    # Recent form header
    row += 1
    ws_rec.merge_cells(f"A{row}:G{row}")
    _rec_cell(row, 1, "RECENT FORM (Last 7 Days)", fill=PatternFill("solid", fgColor="2E4057"),
              font=Font(bold=True, color="FFFFFF", size=11), align=center)
    row += 1
    for col, h in enumerate(["Category", "Win Rate (7d)", "", "", "", "", ""], 1):
        _rec_cell(row, col, h, fill=navy, font=bold_white, align=center)
    row += 1
    for label, key in [("Moneyline", "moneyline"), ("Spread", "spread"), ("Over/Under", "total")]:
        pct = record["recent_7d"][key]
        _rec_cell(row, 1, label, align=left)
        _rec_cell(row, 2, f"{pct}%", align=center)
        row += 1

    # Calibration section
    row += 1
    ws_rec.merge_cells(f"A{row}:G{row}")
    _rec_cell(row, 1, "MODEL CALIBRATION (Auto-Learned Bias Corrections)",
              fill=PatternFill("solid", fgColor="2E4057"),
              font=Font(bold=True, color="FFFFFF", size=11), align=center)
    row += 1
    for col, h in enumerate(["Bias Type", "Value", "Interpretation", "", "", "", ""], 1):
        _rec_cell(row, col, h, fill=navy, font=bold_white, align=center)
    row += 1
    biases = [
        ("Total Projection Bias",      calib.get("total_bias", 0),       "pts model projects too high/low on totals"),
        ("Home Score Bias",            calib.get("home_score_bias", 0),  "pts model over/under-projects home teams"),
        ("Away Score Bias",            calib.get("away_score_bias", 0),  "pts model over/under-projects away teams"),
    ]
    for bname, bval, binterp in biases:
        sign = "+" if bval >= 0 else ""
        _rec_cell(row, 1, bname, align=left)
        _rec_cell(row, 2, f"{sign}{bval:.1f}", align=center,
                  font=Font(color="9C0006" if abs(bval) >= 3 else "375623", bold=abs(bval) >= 3))
        _rec_cell(row, 3, binterp, align=left, font=Font(italic=True, color="595959"))
        row += 1
    row += 1
    for note in calib.get("notes", []):
        ws_rec.merge_cells(f"A{row}:G{row}")
        _rec_cell(row, 1, f"📌 {note}", font=Font(italic=True, color="595959"), align=left)
        row += 1

    _col_widths(ws_rec, [18, 12, 12, 10, 12, 14, 10])

    # ── Sheet 2: Daily Summary ─────────────────────────────────────────────────
    ws_sum = _make_sheet("Summary", [
        "Date", "ML Correct", "ML Total", "ML %",
        "Spread Correct", "Spread Total", "Spread %",
        "O/U Correct", "O/U Total", "O/U %", "Overall %",
    ])
    _col_widths(ws_sum, [12, 11, 9, 8, 14, 12, 10, 11, 9, 8, 10])

    for date in sorted(history.keys()):
        rec = history[date]
        if not rec.get("graded_at"):
            continue
        s  = rec.get("summary", {})
        ml = s.get("moneyline", {})
        sp = s.get("spread", {})
        ou = s.get("total", {})
        all_c = ml.get("correct",0) + sp.get("correct",0) + ou.get("correct",0)
        all_t = ml.get("total",0)   + sp.get("total",0)   + ou.get("total",0)
        ws_sum.append([
            date,
            ml.get("correct",""), ml.get("total",""), ml.get("pct",""),
            sp.get("correct",""), sp.get("total",""), sp.get("pct",""),
            ou.get("correct",""), ou.get("total",""), ou.get("pct",""),
            round(all_c/all_t*100,1) if all_t else "",
        ])

    # ── Sheets 3–5: Per-game detail ────────────────────────────────────────────
    ws_ml = _make_sheet("Moneyline", [
        "Date", "Game", "Pick", "Confidence",
        "Model Home", "Model Away", "Actual Home", "Actual Away",
        "Winner", "Correct",
    ])
    _col_widths(ws_ml, [12, 14, 8, 12, 12, 11, 12, 11, 10, 9])

    ws_sp = _make_sheet("Spread", [
        "Date", "Game", "Pick", "Book Line", "Model Line", "Edge",
        "Actual Home", "Actual Away", "Result", "Correct",
    ])
    _col_widths(ws_sp, [12, 14, 10, 10, 11, 8, 12, 11, 10, 9])

    ws_ou = _make_sheet("Total Over-Under", [
        "Date", "Game", "Pick", "Book Line", "Model Total", "Edge",
        "Actual Total", "Notes", "Correct",
    ])
    _col_widths(ws_ou, [12, 14, 8, 10, 12, 8, 13, 28, 9])

    for date in sorted(history.keys()):
        for g in history[date].get("games", []):
            if not g.get("actuals"):
                continue
            label  = f"{g.get('away','?')} @ {g.get('home','?')}"
            preds  = g.get("predictions", {})
            actual = g.get("actuals", {})
            grades = g.get("grades", {})
            errs   = g.get("error_analysis", {})
            hs     = actual.get("home_score", "")
            as_    = actual.get("away_score", "")

            # Moneyline
            ml_cor = grades.get("winner_correct")
            ws_ml.append([
                date, label,
                preds.get("winner_pick",""), preds.get("winner_confidence",""),
                preds.get("model_home_score",""), preds.get("model_away_score",""),
                hs, as_, actual.get("winner",""),
                "✓" if ml_cor is True else ("✗" if ml_cor is False else "—"),
            ])
            _fill_row(ws_ml, ml_cor)

            # Spread
            sp_cor   = grades.get("spread_correct")
            sp_pick  = f"{preds.get('spread_pick','')} {preds.get('spread_team','')}".strip()
            sp_line  = preds.get("spread_line", "")
            mh, ma   = preds.get("model_home_score"), preds.get("model_away_score")
            model_mar = round(float(mh)-float(ma),1) if mh and ma else ""
            sp_edge  = ""
            try:
                sp_edge = round(float(model_mar) - float(sp_line), 1) if sp_line != "" else ""
            except Exception:
                pass
            ws_sp.append([
                date, label, sp_pick, sp_line, model_mar, sp_edge,
                hs, as_, f"{hs}-{as_}" if hs != "" else "",
                "✓" if sp_cor is True else ("✗" if sp_cor is False else "—"),
            ])
            _fill_row(ws_sp, sp_cor)

            # O/U
            ou_cor   = grades.get("total_correct")
            ou_line  = preds.get("total_line","")
            ou_model = preds.get("model_total","")
            ou_edge  = ""
            try:
                ou_edge = round(float(ou_model)-float(ou_line),1) if ou_line != "" else ""
            except Exception:
                pass
            ou_actual = actual.get("total","")
            # Include auto-learned note about why the pick was wrong
            ou_note = errs.get("missed_total_reason","") or ""
            ws_ou.append([
                date, label,
                preds.get("total_pick",""), ou_line, ou_model, ou_edge,
                ou_actual, ou_note,
                "✓" if ou_cor is True else ("✗" if ou_cor is False else "—"),
            ])
            _fill_row(ws_ou, ou_cor)

    # ── Props Sheets ───────────────────────────────────────────────────────────
    props_history = _load_props_history()

    # Sheet: Props Record (dashboard)
    ws_prec = wb.create_sheet("Props Record")
    ws_prec.sheet_view.showGridLines = False
    prec = get_props_record()

    ws_prec.merge_cells("A1:G1")
    c = ws_prec.cell(1, 1, "🎯  NBA PREDICTOR — PLAYER PROPS RECORD")
    c.fill = navy; c.font = bold_gold; c.alignment = center
    ws_prec.row_dimensions[1].height = 30

    ws_prec.merge_cells("A2:G2")
    c = ws_prec.cell(2, 1, f"Last updated: {datetime.now().strftime('%A %B %d, %Y  %I:%M %p')}")
    c.font = Font(italic=True, color="595959"); c.alignment = center

    # Overall record
    ws_prec.merge_cells("A4:G4")
    c = ws_prec.cell(4, 1, "ALL-TIME PROPS RECORD")
    c.fill = PatternFill("solid", fgColor="2E4057")
    c.font = Font(bold=True, color="FFFFFF", size=11); c.alignment = center

    for col, h in enumerate(["Category", "Hit ✓", "Miss ✗", "Total", "Hit Rate", "Status", ""], 1):
        c = ws_prec.cell(5, col, h)
        c.fill = navy; c.font = bold_white; c.alignment = center

    pct = prec["pct"]
    status = "🔥 HOT" if pct >= 60 else ("✅ OK" if pct >= 50 else "⚠️ COLD")
    for col, val in enumerate(["All Props", prec["hit"], prec["miss"], prec["total"], f"{pct}%", status, ""], 1):
        c = ws_prec.cell(6, col, val)
        c.fill = gold; c.font = Font(bold=True); c.alignment = center

    # By-stat breakdown
    ws_prec.merge_cells("A8:G8")
    c = ws_prec.cell(8, 1, "BREAKDOWN BY STAT")
    c.fill = PatternFill("solid", fgColor="2E4057")
    c.font = Font(bold=True, color="FFFFFF", size=11); c.alignment = center

    for col, h in enumerate(["Stat", "Hit ✓", "Total", "Hit Rate", "", "", ""], 1):
        c = ws_prec.cell(9, col, h)
        c.fill = navy; c.font = bold_white; c.alignment = center

    row_p = 10
    for stat, data in sorted(prec["by_stat"].items()):
        sp = data["pct"]
        row_fill = PatternFill("solid", fgColor="C6EFCE") if sp >= 55 else PatternFill("solid", fgColor="FFC7CE")
        for col, val in enumerate([stat, data["hit"], data["total"], f"{sp}%", "", "", ""], 1):
            c = ws_prec.cell(row_p, col, val)
            c.fill = row_fill; c.alignment = center
        row_p += 1

    # Recent 7d
    row_p += 1
    ws_prec.merge_cells(f"A{row_p}:G{row_p}")
    c = ws_prec.cell(row_p, 1, f"Recent 7-Day Hit Rate: {prec['recent_7d']}%   |   Days Graded: {prec['days_graded']}")
    c.font = Font(bold=True, italic=True); c.alignment = center

    _col_widths(ws_prec, [18, 10, 10, 10, 12, 14, 10])

    # Sheet: Props Summary (one row per date)
    ws_psum = _make_sheet("Props Summary", [
        "Date", "Hit", "Miss", "Total", "Hit Rate %",
        "PTS %", "REB %", "AST %", "FG3M %", "Combos %",
    ])
    _col_widths(ws_psum, [12, 8, 8, 8, 12, 9, 9, 9, 9, 10])

    for date in sorted(props_history.keys()):
        rec = props_history[date]
        if not rec.get("graded_at"):
            continue
        s = rec.get("summary", {})
        bs = s.get("by_stat", {})

        def _bs(key):
            return f"{bs[key]['pct']}%" if key in bs else ""

        combo_keys = [k for k in bs if "+" in k]
        combo_hit = sum(bs[k]["hit"] for k in combo_keys)
        combo_tot = sum(bs[k]["total"] for k in combo_keys)
        combo_pct = f"{_pct(combo_hit, combo_tot)}%" if combo_tot else ""

        ws_psum.append([
            date, s.get("hit",""), s.get("miss",""), s.get("total",""),
            f"{s.get('pct','')}%",
            _bs("PTS"), _bs("REB"), _bs("AST"), _bs("FG3M"), combo_pct,
        ])

    # Sheet: Props Detail (one row per prop)
    ws_pdet = _make_sheet("Props Detail", [
        "Date", "Player", "Team", "Opponent", "Stat", "Direction",
        "Line", "Avg (L10)", "Hit Rate", "EV", "Actual", "Hit?",
    ])
    _col_widths(ws_pdet, [12, 22, 8, 10, 14, 10, 8, 10, 10, 8, 8, 8])

    for date in sorted(props_history.keys()):
        for prop in props_history[date].get("props", []):
            if prop.get("actual") is None:
                continue   # not yet graded
            hit_val = prop.get("hit")
            ws_pdet.append([
                date,
                prop.get("player", ""),
                prop.get("team", ""),
                prop.get("opponent", ""),
                prop.get("stat", ""),
                prop.get("direction", ""),
                prop.get("line", ""),
                prop.get("avg", ""),
                f"{round(prop.get('hit_rate',0)*100,1)}%",
                round(prop.get("ev", 0), 3),
                prop.get("actual", ""),
                "✓" if hit_val is True else ("✗" if hit_val is False else "—"),
            ])
            _fill_row(ws_pdet, hit_val)

    wb.save(path)
    print(f"[Tracker] Excel exported → {path}")
    return str(path)


# ─── ESPN score fetcher ────────────────────────────────────────────────────────

def _fetch_actuals(date: str) -> dict[str, dict]:
    """Fetch final scores from ESPN for a given date."""
    try:
        import requests
        from utils.data_fetch import _normalize_espn_abbrev  # type: ignore
        url  = "https://site.api.espn.com/apis/site/v2/sports/basketball/nba/scoreboard"
        resp = requests.get(url, params={"dates": date.replace("-", "")}, timeout=10)
        resp.raise_for_status()
        data = resp.json()

        results: dict[str, dict] = {}
        for event in data.get("events", []):
            if "final" not in event.get("status",{}).get("type",{}).get("description","").lower():
                continue
            comp = event["competitions"][0]
            hs = as_ = None
            ha = aa = ""
            for competitor in comp["competitors"]:
                abbr  = _normalize_espn_abbrev(competitor["team"]["abbreviation"])
                score = int(competitor.get("score", 0))
                if competitor["homeAway"] == "home":
                    ha = abbr; hs = score
                else:
                    aa = abbr; as_ = score
            if hs is not None and as_ is not None:
                results[f"{aa}@{ha}"] = {"home_score": hs, "away_score": as_}
        return results
    except Exception as e:
        print(f"[Tracker] Error fetching actuals for {date}: {e}")
        return {}
